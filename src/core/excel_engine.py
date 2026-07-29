from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.core.document import DocumentABC
from src.rendering.styles import TextStyle


class ExcelEngine(DocumentABC):

    def __init__(self, path: str | Path | None = None) -> None:
        super().__init__(path)
        self._wb: Workbook | None = None
        self._base_style: TextStyle = TextStyle.default_excel()

    @property
    def wb(self) -> Workbook:
        if self._wb is None:
            raise RuntimeError('No workbook loaded. Call create() or open() first.')
        return self._wb

    def create(self) -> None:
        self._wb = Workbook()
        self._path = None
        self._base_style = TextStyle.default_excel()

    def open(self, path: str | Path) -> None:
        self._path = Path(path)
        if not self._path.exists():
            raise FileNotFoundError(f'File not found: {self._path}')
        self._wb = load_workbook(str(self._path))
        self._base_style = TextStyle.default_excel()

    def save(self, path: str | Path | None = None) -> None:
        if path is not None:
            self._path = Path(path)
        if self._path is None:
            raise ValueError('No output path specified.')
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(self._path))

    def get_base_style(self) -> TextStyle:
        return self._base_style

    def add_text(
        self,
        text: str,
        bold: bool = False,
        italic: bool = False,
        font_name: str | None = None,
        font_size: int | None = None,
        color: str | None = None,
        alignment: str | None = None,
        text_style: TextStyle | None = None,
        **kwargs: Any,
    ) -> Any:
        column = int(kwargs.get('column', 1))
        inline = TextStyle(
            bold=bold or None,
            italic=italic or None,
            font_name=font_name,
            font_size=font_size,
            color=color,
            alignment=alignment,
        )
        final = self._base_style
        if text_style is not None:
            final = final.merge(text_style)
        final = final.merge(inline)

        ws = self.wb.active
        if ws.max_row is None or (ws.max_row == 1 and ws['A1'].value is None):
            row = 1
        else:
            row = ws.max_row + 1

        last_cell = None
        for line in text.split('\n'):
            cell = ws.cell(row=row, column=column, value=line)
            font_kwargs: dict[str, Any] = {}
            if final.font_name:
                font_kwargs['name'] = final.font_name
            if final.font_size:
                font_kwargs['size'] = final.font_size
            if final.bold is not None:
                font_kwargs['bold'] = final.bold
            if final.italic is not None:
                font_kwargs['italic'] = final.italic
            if final.color:
                font_kwargs['color'] = final.color.lstrip('#')
            cell.font = Font(**font_kwargs)
            if final.alignment:
                align_map = {
                    'left': 'left', 'center': 'center',
                    'right': 'right', 'justify': 'justify',
                }
                cell.alignment = Alignment(horizontal=align_map.get(final.alignment, 'left'))
            last_cell = cell
            row += 1

        return last_cell

    def add_styled_content(
        self,
        tokens: list[dict[str, Any]],
    ) -> Any:
        current_style = self._base_style
        results: list[Any] = []

        for token in tokens:
            token_type = token.get('type', 'text')
            content = token.get('content', '')

            if token_type == 'text':
                results.append(self.add_text(str(content), text_style=current_style))
            elif token_type == 'command':
                cmd = token.get('command', '')
                if cmd in ('newpage', 'heading', 'includegraphics'):
                    continue
                token_style = TextStyle.from_latex_token(token)
                merged = current_style.merge(token_style)
                inner = token.get('content', '')
                if inner:
                    results.append(self.add_text(inner, text_style=merged))

        return results

    def add_latex_content(self, text: str) -> Any:
        from src.rendering.latex_parser import parse_structured
        tokens = parse_structured(text)
        return self.add_styled_content(tokens)

    def replace_text(self, old: str, new: str, regex: bool = False) -> int:
        import re as _re
        count = 0
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    if regex:
                        new_val = _re.sub(old, new, str(cell.value))
                        if new_val != str(cell.value):
                            cell.value = new_val
                            count += 1
                    else:
                        if old in str(cell.value):
                            cell.value = str(cell.value).replace(old, new)
                            count += 1
        return count

    def set_style(self, style_str: str) -> None:
        new_style = TextStyle.from_string(style_str)
        self._base_style = self._base_style.merge(new_style)

    def apply_style_to_all(self) -> None:
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        font_kwargs: dict[str, Any] = {}
                        if self._base_style.font_name:
                            font_kwargs['name'] = self._base_style.font_name
                        if self._base_style.font_size:
                            font_kwargs['size'] = self._base_style.font_size
                        if self._base_style.bold is not None:
                            font_kwargs['bold'] = self._base_style.bold
                        if self._base_style.italic is not None:
                            font_kwargs['italic'] = self._base_style.italic
                        if self._base_style.color:
                            font_kwargs['color'] = self._base_style.color.lstrip('#')
                        if font_kwargs:
                            cell.font = Font(**font_kwargs)

    def to_pdf(self, output_path: str | Path) -> None:
        self.save()
        from src.transform.pdf import excel_to_pdf
        excel_to_pdf(str(self._path), str(output_path))

    def add_sheet(self, name: str) -> Any:
        return self.wb.create_sheet(title=name)

    def delete_sheet(self, name: str) -> None:
        if name in self.wb.sheetnames:
            del self.wb[name]

    def rename_sheet(self, old_name: str, new_name: str) -> None:
        if old_name in self.wb.sheetnames:
            self.wb[old_name].title = new_name

    def set_column_width(self, col_index: int, width: float) -> None:
        ws = self.wb.active
        ws.column_dimensions[get_column_letter(col_index)].width = width

    def set_row_height(self, row_index: int, height: float) -> None:
        ws = self.wb.active
        ws.row_dimensions[row_index].height = height

    def set_formula(self, cell_ref: str, formula: str) -> None:
        ws = self.wb.active
        ws[cell_ref] = formula

    def import_csv(self, csv_path: str | Path) -> None:
        import csv
        ws = self.wb.active
        with open(csv_path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row_idx, row_data in enumerate(reader, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

    def export_csv(self, output_path: str | Path) -> None:
        import csv
        ws = self.wb.active
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

    def get_metadata(self) -> dict[str, str | None]:
        return {
            'author': self.wb.properties.creator,
            'title': self.wb.properties.title,
            'subject': self.wb.properties.subject,
            'category': self.wb.properties.category,
            'keywords': self.wb.properties.keywords,
            'comments': self.wb.properties.description,
        }

    def set_metadata(self, **kwargs: str) -> None:
        props = self.wb.properties
        for key, value in kwargs.items():
            key_lower = key.lower()
            if key_lower == 'author':
                props.creator = value
            elif key_lower == 'title':
                props.title = value
            elif key_lower == 'subject':
                props.subject = value
            elif key_lower == 'category':
                props.category = value
            elif key_lower == 'keywords':
                props.keywords = value
            elif key_lower == 'comments':
                props.description = value

    def add_comment(self, cell_ref: str, text: str) -> None:
        from openpyxl.comments import Comment
        ws = self.wb.active
        ws[cell_ref].comment = Comment(text, 'TianshangScribe')

    def set_protection(self, password: str) -> None:
        self.wb.security.workbook_password = password

    def unprotect(self) -> None:
        self.wb.security.workbook_password = ''

    def export_json(self, output_path: str | Path) -> None:
        import json
        ws = self.wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump([], f)
            return
        headers = [str(c) if c is not None else f'col_{i + 1}' for i, c in enumerate(rows[0])]
        data = [
            {headers[j]: str(c) if c is not None else '' for j, c in enumerate(r)}
            for r in rows[1:]
        ]
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def export_html(self, output_path: str | Path) -> None:
        ws = self.wb.active
        rows = list(ws.iter_rows(values_only=True))
        html_parts = [
            '<!DOCTYPE html><html><head><meta charset="utf-8"><title>',
            self.wb.properties.title or 'Excel Export',
            '</title><style>table{border-collapse:collapse}',
            'td,th{border:1px solid #ccc;padding:4px 8px}</style>',
            '</head><body><table>',
        ]
        first = True
        for row in rows:
            html_parts.append('<tr>')
            for cell in row:
                tag = 'th' if first else 'td'
                html_parts.append(f'<{tag}>{cell or ""}</{tag}>')
            html_parts.append('</tr>')
            first = False
        html_parts.append('</table></body></html>')
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(''.join(html_parts))

    def sort(self, cell_range: str, order: str = 'asc') -> None:
        import re
        ws = self.wb.active
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', cell_range)
        if not match:
            raise ValueError(f'Invalid cell range: {cell_range}')
        c1, r1_s, c2, r2_s = match.groups()
        r1 = int(r1_s)
        r2 = int(r2_s)
        data = []
        for row in range(r1, r2 + 1):
            val = ws[f'{c1}{row}'].value
            data.append((val, row))
        data.sort(reverse=(order == 'desc'), key=lambda x: (x[0] is None, x[0] or ''))
        for i, (val, _orig_row) in enumerate(data, start=r1):
            ws[f'{c1}{i}'].value = val

    def add_chart(self, chart_type: str, data_range: str, position: str = 'E2') -> None:
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        ws = self.wb.active
        chart_classes = {'bar': BarChart, 'line': LineChart, 'pie': PieChart}
        chart_cls = chart_classes.get(chart_type)
        if chart_cls is None:
            raise ValueError(
                f'Unsupported chart type: {chart_type}. Use bar, line, or pie.'
            )
        chart = chart_cls()
        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, position)

    def merge_workbooks(self, paths: list[str]) -> None:
        for p in paths:
            src = load_workbook(p)
            for sheet in src.worksheets:
                new_sheet = self.wb.create_sheet(title=sheet.title)
                for row in sheet.iter_rows():
                    for cell in row:
                        new_sheet[cell.coordinate] = cell.value

    def split_by_sheet(self, output_dir: str | Path) -> list[Path]:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        results: list[Path] = []
        for sheet in self.wb.worksheets:
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet.title
            for row in sheet.iter_rows():
                for cell in row:
                    new_ws[cell.coordinate] = cell.value
            out_path = output_dir / f'{sheet.title}.xlsx'
            new_wb.save(str(out_path))
            results.append(out_path)
        return results

    def clear_content(self) -> None:
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
