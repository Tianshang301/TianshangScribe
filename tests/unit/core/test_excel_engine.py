"""Unit tests for the Excel engine."""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

from tianshang_scribe.core.document import open_document
from tianshang_scribe.core.excel_engine import ExcelEngine


class TestExcelEngine:
    @pytest.fixture
    def engine(self) -> ExcelEngine:
        e = ExcelEngine()
        e.create()
        return e

    def test_create_empty(self, engine: ExcelEngine) -> None:
        ws = engine.wb.active
        assert ws is not None

    def test_add_text(self, engine: ExcelEngine) -> None:
        engine.add_text('Cell Content')
        ws = engine.wb.active
        assert ws.cell(row=1, column=1).value == 'Cell Content'

    def test_save_and_reopen(self, engine: ExcelEngine) -> None:
        engine.add_text('Excel Test')
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / 'test.xlsx'
            engine.save(path)
            assert path.exists()
            reopened = open_document(path)
            assert isinstance(reopened, ExcelEngine)
            ws = reopened.wb.active
            assert ws.cell(row=1, column=1).value == 'Excel Test'

    def test_add_text_with_column(self, engine: ExcelEngine) -> None:
        engine.add_text('Col2', column=2)
        ws = engine.wb.active
        assert ws.cell(row=1, column=2).value == 'Col2'

    def test_add_text_multiline(self, engine: ExcelEngine) -> None:
        engine.add_text('A\nB\nC')
        ws = engine.wb.active
        assert ws.cell(row=1, column=1).value == 'A'
        assert ws.cell(row=2, column=1).value == 'B'
        assert ws.cell(row=3, column=1).value == 'C'

    def test_replace_text_exact(self, engine: ExcelEngine) -> None:
        engine.add_text('Hello World')
        count = engine.replace_text('World', 'Earth')
        assert count >= 1
        assert engine.wb.active.cell(row=1, column=1).value == 'Hello Earth'

    def test_get_metadata(self, engine: ExcelEngine) -> None:
        metadata = engine.get_metadata()
        assert 'author' in metadata
        assert 'title' in metadata

    def test_set_metadata(self, engine: ExcelEngine) -> None:
        engine.set_metadata(author='Test Author', title='Test')
        metadata = engine.get_metadata()
        assert metadata['author'] == 'Test Author'
        assert metadata['title'] == 'Test'

    def test_set_style(self, engine: ExcelEngine) -> None:
        engine.set_style('font=Courier New,size=14')
        assert engine._base_style.font_name == 'Courier New'
        assert engine._base_style.font_size == 14

    def test_add_sheet(self, engine: ExcelEngine) -> None:
        count_before = len(engine.wb.sheetnames)
        engine.add_sheet('NewSheet')
        assert len(engine.wb.sheetnames) == count_before + 1
        assert 'NewSheet' in engine.wb.sheetnames

    def test_delete_sheet(self, engine: ExcelEngine) -> None:
        engine.add_sheet('ToDelete')
        engine.delete_sheet('ToDelete')
        assert 'ToDelete' not in engine.wb.sheetnames

    def test_rename_sheet(self, engine: ExcelEngine) -> None:
        engine.add_sheet('OldName')
        engine.rename_sheet('OldName', 'NewName')
        assert 'NewName' in engine.wb.sheetnames
        assert 'OldName' not in engine.wb.sheetnames

    def test_set_column_width(self, engine: ExcelEngine) -> None:
        engine.set_column_width(2, 25.5)
        ws = engine.wb.active
        assert ws.column_dimensions['B'].width == 25.5

    def test_set_row_height(self, engine: ExcelEngine) -> None:
        engine.set_row_height(3, 30.0)
        ws = engine.wb.active
        assert ws.row_dimensions[3].height == 30.0

    def test_set_formula(self, engine: ExcelEngine) -> None:
        engine.set_formula('A1', '=SUM(B1:B10)')
        ws = engine.wb.active
        assert ws['A1'].value == '=SUM(B1:B10)'

    def test_import_csv(self, engine: ExcelEngine) -> None:
        import csv

        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / 'test.csv'
            with open(csv_path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Name', 'Age'])
                writer.writerow(['Alice', '30'])
            engine.import_csv(str(csv_path))
            ws = engine.wb.active
            assert ws.cell(row=1, column=1).value == 'Name'
            assert ws.cell(row=2, column=2).value == '30'

    def test_export_csv(self, engine: ExcelEngine) -> None:
        engine.add_text('Name', column=1)
        engine.add_text('Alice', column=1)
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / 'out.csv'
            engine.export_csv(str(out_path))
            assert out_path.exists()
            content = out_path.read_text()
            assert 'Name' in content
            assert 'Alice' in content

    def test_export_json(self, engine: ExcelEngine) -> None:
        engine.add_text('Name', column=1)
        engine.add_text('Alice', column=1)
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / 'out.json'
            engine.export_json(str(out_path))
            assert out_path.exists()
            import json

            data = json.loads(out_path.read_text())
            assert isinstance(data, list)
            assert data[0]['Name'] == 'Alice'

    def test_export_html(self, engine: ExcelEngine) -> None:
        engine.add_text('Header', column=1)
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / 'out.html'
            engine.export_html(str(out_path))
            assert out_path.exists()
            content = out_path.read_text()
            assert '<table>' in content
            assert 'Header' in content

    def test_sort_asc(self, engine: ExcelEngine) -> None:
        engine.add_text('B', column=1)
        engine.add_text('A', column=1)
        engine.add_text('C', column=1)
        engine.sort('A1:A3', 'asc')
        ws = engine.wb.active
        assert ws.cell(row=1, column=1).value == 'A'
        assert ws.cell(row=2, column=1).value == 'B'
        assert ws.cell(row=3, column=1).value == 'C'

    def test_sort_desc(self, engine: ExcelEngine) -> None:
        engine.add_text('A', column=1)
        engine.add_text('B', column=1)
        engine.add_text('C', column=1)
        engine.sort('A1:A3', 'desc')
        ws = engine.wb.active
        assert ws.cell(row=1, column=1).value == 'C'
        assert ws.cell(row=3, column=1).value == 'A'

    def test_add_chart_bar(self, engine: ExcelEngine) -> None:
        ws = engine.wb.active
        engine.add_text('Cats', column=1)
        engine.add_text('10', column=1)
        engine.add_text('Dogs', column=2)
        engine.add_text('20', column=2)
        sheet_name = ws.title
        engine.add_chart('bar', f"'{sheet_name}'!A1:B2")
        assert len(ws._charts) >= 1

    # ---- Step 1 (E1): freeze panes ----
    def test_freeze_panes_baseline(self, engine: ExcelEngine, tmp_path: Path) -> None:
        engine.add_text('hdr', column=1)
        engine.freeze_panes('A2')
        assert engine.wb.active.freeze_panes == 'A2'
        out = tmp_path / 'frozen.xlsx'
        engine.save(out)
        reopened = ExcelEngine()
        reopened.open(str(out))
        assert reopened.wb.active.freeze_panes == 'A2'

    # ---- Step 2 (E2): number format ----
    def test_set_number_format_baseline(self, engine: ExcelEngine, tmp_path: Path) -> None:
        ws = engine.wb.active
        ws['A1'] = 0.1234
        ws['A2'] = 0.5678
        engine.set_number_format('A1:A2', '0.00%')
        assert ws['A1'].number_format == '0.00%'
        assert ws['A2'].number_format == '0.00%'
        engine.set_number_format('B1', 'yyyy-mm-dd')
        assert ws['B1'].number_format == 'yyyy-mm-dd'
        out = tmp_path / 'fmt.xlsx'
        engine.save(out)
        reopened = ExcelEngine()
        reopened.open(str(out))
        assert reopened.wb.active['A1'].number_format == '0.00%'

    def test_set_number_format_invalid_range_raises(self, engine: ExcelEngine) -> None:
        with pytest.raises(ValueError, match='Invalid cell range'):
            engine.set_number_format('not-a-range', '0.00')

    # ---- Step 3 (E4): conditional formatting ----
    def test_add_conditional_format_baseline(self, engine: ExcelEngine, tmp_path: Path) -> None:
        ws = engine.wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i * 10)
        engine.add_conditional_format('A1:A5', 'color_scale')
        engine.add_conditional_format('B1:B5', 'data_bar')
        engine.add_conditional_format('C1:C5', 'cell_is', operator='greaterThan', formula='20')
        out = tmp_path / 'cf.xlsx'
        engine.save(out)
        reopened = ExcelEngine()
        reopened.open(str(out))
        rules = list(reopened.wb.active.conditional_formatting)
        assert len(rules) >= 3

    def test_add_conditional_format_unsupported_raises(self, engine: ExcelEngine) -> None:
        with pytest.raises(ValueError, match='Unsupported conditional format type'):
            engine.add_conditional_format('A1:A5', 'bogus')

    # ---- Step 4 (E5): data validation ----
    def test_add_data_validation_baseline(self, engine: ExcelEngine, tmp_path: Path) -> None:
        engine.add_data_validation('A1:A10', 'list', 'yes,no')
        engine.add_data_validation('B1:B10', 'whole', '1', '100')
        out = tmp_path / 'dv.xlsx'
        engine.save(out)
        reopened = ExcelEngine()
        reopened.open(str(out))
        dvs = reopened.wb.active.data_validations.dataValidation
        types = {dv.type for dv in dvs}
        assert 'list' in types
        assert 'whole' in types

    def test_add_data_validation_list_quotes(self, engine: ExcelEngine) -> None:
        engine.add_data_validation('A1:A5', 'list', 'x,y,z')
        dv = engine.wb.active.data_validations.dataValidation[0]
        assert dv.formula1 == '"x,y,z"'

    # ---- Step 5 (E3): range style (border/fill) ----
    def test_set_range_style_baseline(self, engine: ExcelEngine, tmp_path: Path) -> None:
        engine.set_range_style('B2:C3', 'border=medium,fill=FFFF00')
        out = tmp_path / 'style.xlsx'
        engine.save(out)
        reopened = ExcelEngine()
        reopened.open(str(out))
        cell = reopened.wb.active['B2']
        assert cell.border.left is not None
        assert 'FFFF00' in cell.fill.fgColor.rgb

    def test_set_range_style_invalid_range_raises(self, engine: ExcelEngine) -> None:
        with pytest.raises(ValueError, match='Invalid cell range'):
            engine.set_range_style('bad', 'border=thin')

    def test_set_range_style_font_alignment_number_format(
        self, engine: ExcelEngine, tmp_path: Path
    ) -> None:
        engine.set_range_style(
            'B2:C3', 'font=Times,size=14,bold,italic,color=FF0000,align=center,numfmt=0.00%'
        )
        engine.wb.active['B2'] = 0.5
        out = tmp_path / 'style2.xlsx'
        engine.save(out)
        reopened = ExcelEngine()
        reopened.open(str(out))
        cell = reopened.wb.active['B2']
        assert cell.font.bold is True
        assert cell.font.italic is True
        assert cell.font.size == 14
        assert cell.font.name == 'Times'
        assert 'FF0000' in str(cell.font.color.rgb)
        assert cell.alignment.horizontal == 'center'
        assert cell.number_format == '0.00%'

    # ---- Step 7 (E7/E8/E9): hyperlink, named range, auto_fit ----
    def test_add_hyperlink_baseline(self, engine: ExcelEngine, tmp_path: Path) -> None:
        engine.add_hyperlink('A1', 'https://example.com')
        assert engine.wb.active['A1'].hyperlink.target == 'https://example.com'
        out = tmp_path / 'link.xlsx'
        engine.save(out)
        reopened = ExcelEngine()
        reopened.open(str(out))
        assert reopened.wb.active['A1'].hyperlink.target == 'https://example.com'

    def test_set_named_range_baseline(self, engine: ExcelEngine, tmp_path: Path) -> None:
        engine.set_named_range('MyRange', 'A1:B2')
        assert 'MyRange' in engine.wb.defined_names
        out = tmp_path / 'named.xlsx'
        engine.save(out)
        reopened = ExcelEngine()
        reopened.open(str(out))
        assert 'MyRange' in reopened.wb.defined_names

    def test_auto_fit_baseline(self, engine: ExcelEngine) -> None:
        engine.add_text('a very long label here', column=1)
        engine.auto_fit(1)
        width = engine.wb.active.column_dimensions['A'].width
        assert width is not None and width > 8

    def test_add_chart_unsupported_raises(self, engine: ExcelEngine) -> None:
        with pytest.raises(ValueError, match='Unsupported chart type'):
            engine.add_chart('bogus', 'Sheet!A1:B2')

    # ---- Step 6 (E6): chart type extension ----
    @pytest.mark.parametrize('chart_type', ['area', 'doughnut', 'scatter'])
    def test_add_chart_extended_types_baseline(self, engine: ExcelEngine, chart_type: str) -> None:
        ws = engine.wb.active
        sheet = ws.title
        ws['A1'] = 'Cat'
        ws['A2'] = 'X'
        ws['A3'] = 'Y'
        ws['B1'] = 'Val'
        ws['B2'] = 10
        ws['B3'] = 20
        engine.add_chart(chart_type, f'{sheet}!A1:B3')
        assert len(ws._charts) >= 1

    def test_merge_workbooks(self, engine: ExcelEngine) -> None:
        engine.add_sheet('Data')
        engine.add_text('Hello', column=1)
        with tempfile.TemporaryDirectory() as tmpdir:
            p1 = Path(tmpdir) / 'src.xlsx'
            engine.save(str(p1))
            e2 = ExcelEngine()
            e2.create()
            e2.merge_workbooks([str(p1)])
            assert 'Data' in e2.wb.sheetnames

    def test_split_by_sheet(self, engine: ExcelEngine) -> None:
        engine.add_sheet('Sheet1')
        engine.add_sheet('Sheet2')
        with tempfile.TemporaryDirectory() as tmpdir:
            results = engine.split_by_sheet(tmpdir)
            assert len(results) == 3
            for p in results:
                assert p.exists()

    def test_clear_content(self, engine: ExcelEngine) -> None:
        engine.add_text('Data', column=1)
        engine.add_text('More', column=1)
        engine.clear_content()
        ws = engine.wb.active
        assert ws.cell(row=1, column=1).value is None
        assert ws.cell(row=2, column=1).value is None

    def test_set_protection(self, engine: ExcelEngine) -> None:
        engine.set_protection('secret123')
        assert engine.wb.security.workbook_password is not None

    def test_unprotect(self, engine: ExcelEngine) -> None:
        engine.set_protection('secret123')
        engine.unprotect()
        assert engine.wb.security.workbook_password is not None

    def test_add_comment(self, engine: ExcelEngine) -> None:
        engine.add_text('Cell', column=1)
        engine.add_comment('A1', 'A note')
        ws = engine.wb.active
        assert ws['A1'].comment is not None
        assert ws['A1'].comment.text == 'A note'


class TestExcelEdge:
    @pytest.fixture
    def engine(self) -> ExcelEngine:
        e = ExcelEngine()
        e.create()
        return e

    def test_wb_property_unloaded(self) -> None:
        e = ExcelEngine()
        with pytest.raises(RuntimeError):
            _ = e.wb

    def test_open_nonexistent(self) -> None:
        e = ExcelEngine()
        with pytest.raises(FileNotFoundError):
            e.open('nope.xlsx')

    def test_save_without_path(self, engine: ExcelEngine) -> None:
        with pytest.raises(ValueError):
            engine.save()

    def test_get_base_style(self, engine: ExcelEngine) -> None:
        assert engine.get_base_style().font_name == 'Calibri'

    def test_add_text_full_style(self, engine: ExcelEngine) -> None:
        cell = engine.add_text(
            'Styled',
            bold=True,
            italic=True,
            font_name='Arial',
            font_size=14,
            color='FF0000',
            alignment='center',
        )
        assert cell.value == 'Styled'
        assert cell.font.bold is True
        assert cell.font.name == 'Arial'
        assert cell.font.size == 14

    def test_add_text_with_style_object(self, engine: ExcelEngine) -> None:
        from tianshang_scribe.rendering.styles import TextStyle

        engine.add_text('Styled', text_style=TextStyle(font_name='Arial', font_size=14))
        assert engine.wb.active['A1'].font.name == 'Arial'

    def test_add_styled_content(self, engine: ExcelEngine) -> None:
        engine.add_styled_content(
            [
                {'type': 'text', 'content': 'plain'},
                {'type': 'command', 'command': 'newpage'},
                {'type': 'command', 'command': 'bfseries', 'content': 'bold'},
                {'type': 'command', 'command': 'itshape', 'content': ''},
            ]
        )
        assert engine.wb.active['A1'].value == 'plain'
        assert engine.wb.active['A2'].value == 'bold'

    def test_add_latex_content(self, engine: ExcelEngine) -> None:
        engine.add_latex_content(r'\bfseries{Important}')
        assert 'Important' in engine.extract_text()

    def test_replace_text_regex(self, engine: ExcelEngine) -> None:
        engine.add_text('Order 1')
        count = engine.replace_text(r'\d+', '2', regex=True)
        assert count >= 1
        assert engine.wb.active['A1'].value == 'Order 2'

    def test_replace_text_skips_none(self, engine: ExcelEngine) -> None:
        engine.wb.active['A1'] = None
        assert engine.replace_text('x', 'y') == 0

    def test_apply_style_to_all(self, engine: ExcelEngine) -> None:
        engine.add_text('data')
        engine.set_style('font=Courier New,bold')
        engine.apply_style_to_all()
        assert engine.wb.active['A1'].font.name == 'Courier New'
        assert engine.wb.active['A1'].font.bold is True

    def test_delete_sheet_missing(self, engine: ExcelEngine) -> None:
        engine.delete_sheet('NoSuchSheet')

    def test_rename_sheet_missing(self, engine: ExcelEngine) -> None:
        engine.rename_sheet('NoSuchSheet', 'x')

    def test_to_pdf(self, engine: ExcelEngine, tmp_path: Path, monkeypatch) -> None:
        engine.add_text('x')
        out = tmp_path / 'o.pdf'
        calls = []
        monkeypatch.setattr(engine, 'save', lambda *a, **k: calls.append('save'))
        monkeypatch.setattr(
            'tianshang_scribe.transform.pdf.excel_to_pdf', lambda s, d: calls.append((s, d))
        )
        engine.to_pdf(out)
        assert 'save' in calls

    def test_import_json_non_list(self, engine: ExcelEngine, tmp_path: Path) -> None:
        p = tmp_path / 'd.json'
        p.write_text('{"a":1}', encoding='utf-8')
        with pytest.raises(ValueError, match='non-empty'):
            engine.import_json(str(p))

    def test_import_json_scalars(self, engine: ExcelEngine, tmp_path: Path) -> None:
        p = tmp_path / 'd.json'
        p.write_text('[5, 6]', encoding='utf-8')
        engine.import_json(str(p))
        assert engine.wb.active['A1'].value == 5

    def test_export_json_empty(self, engine: ExcelEngine, tmp_path: Path) -> None:
        out = tmp_path / 'o.json'
        engine.export_json(str(out))
        import json

        assert json.loads(out.read_text()) == []

    def test_export_json_with_nones(self, engine: ExcelEngine, tmp_path: Path) -> None:
        ws = engine.wb.active
        ws['A1'] = 'H'
        ws['A2'] = 'v'
        out = tmp_path / 'o.json'
        engine.export_json(str(out))
        import json

        data = json.loads(out.read_text())
        assert data[0]['H'] == 'v'

    def test_set_metadata_all_keys(self, engine: ExcelEngine) -> None:
        engine.set_metadata(
            author='a', title='t', subject='s', category='c', keywords='k', comments='m'
        )
        md = engine.get_metadata()
        assert md['author'] == 'a'
        assert md['title'] == 't'
        assert md['comments'] == 'm'

    def test_add_chart_line_pie(self, engine: ExcelEngine) -> None:
        ws = engine.wb.active
        ws['A1'] = 'x'
        ws['A2'] = '1'
        sheet = ws.title
        engine.add_chart('line', f"'{sheet}'!A1:A2")
        engine.add_chart('pie', f"'{sheet}'!A1:A2")
        assert len(ws._charts) == 2

    def test_sort_invalid_range(self, engine: ExcelEngine) -> None:
        with pytest.raises(ValueError, match='Invalid cell range'):
            engine.sort('badrange')

    def test_sort_desc_with_none(self, engine: ExcelEngine) -> None:
        ws = engine.wb.active
        ws['A1'] = 'a'
        ws['A2'] = 'b'
        ws['A3'] = None
        engine.sort('A1:A3', 'desc')
        assert ws['A1'].value is None
        assert ws['A2'].value == 'b'
        assert ws['A3'].value == 'a'

    def test_sort_preserves_row_integrity(self, engine: ExcelEngine) -> None:
        ws = engine.wb.active
        ws['A1'], ws['B1'] = 2, 'two'
        ws['A2'], ws['B2'] = 1, 'one'
        ws['A3'], ws['B3'] = 3, 'three'
        engine.sort('A1:B3', 'asc')
        assert (ws['A1'].value, ws['B1'].value) == (1, 'one')
        assert (ws['A2'].value, ws['B2'].value) == (2, 'two')
        assert (ws['A3'].value, ws['B3'].value) == (3, 'three')

    def test_sort_multi_column_mixed_types(self, engine: ExcelEngine) -> None:
        ws = engine.wb.active
        # mixed types must not raise and rows stay intact
        ws['A1'], ws['B1'] = 'x', 10
        ws['A2'], ws['B2'] = 'x', 2
        ws['A3'], ws['B3'] = 5, 99
        engine.sort('A1:B3', 'asc', key_columns=[0, 1], orders=['asc', 'desc'])
        # numbers sort before strings; within 'x', larger B comes first on desc
        assert ws['A1'].value == 5
        assert (ws['A2'].value, ws['B2'].value) == ('x', 10)
        assert (ws['A3'].value, ws['B3'].value) == ('x', 2)

    def test_sort_key_columns_out_of_range(self, engine: ExcelEngine) -> None:
        with pytest.raises(ValueError, match='out of range'):
            engine.sort('A1:B3', 'asc', key_columns=[5])

    def test_select_sheet_targets_operations(self, engine: ExcelEngine) -> None:
        engine.add_sheet('Data')
        engine.select_sheet('Data')
        engine.add_text('hello')
        # written to the selected sheet, not the default active sheet
        assert engine.wb['Data']['A1'].value == 'hello'
        assert engine.wb.active['A1'].value is None
        with pytest.raises(ValueError, match='not found'):
            engine.select_sheet('nonexistent')

    def test_clear_formats(self, engine: ExcelEngine) -> None:
        engine.add_text('styled', bold=True)
        engine.clear_formats()
        assert engine.wb.active['A1'].font.bold is not True

    def test_clear_links(self, engine: ExcelEngine) -> None:
        engine.add_text('x')
        engine.clear_links()

    def test_extract_text_no_data(self, engine: ExcelEngine) -> None:
        assert engine.extract_text() == ''

    def test_extract_tables_empty(self, engine: ExcelEngine) -> None:
        assert engine.extract_tables() == []

    def test_extract_images_no_images(self, engine: ExcelEngine, tmp_path: Path) -> None:
        saved = engine.extract_images(str(tmp_path / 'imgs'))
        assert saved == []

    def test_extract_images_with_image(self, engine: ExcelEngine, tmp_path: Path) -> None:
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image

        img = tmp_path / 'pic.png'
        Image.new('RGB', (16, 16), (255, 0, 0)).save(img)
        ws = engine.wb.active
        ws.add_image(XLImage(str(img)), 'A1')
        saved = engine.extract_images(str(tmp_path / 'out'))
        assert len(saved) == 1


class TestExcelImportJson:
    def test_import_json_dicts(self, tmp_path: Path) -> None:
        data = [{'name': 'Alice', 'age': 30}, {'name': 'Bob', 'age': 25}]
        json_path = tmp_path / 'data.json'
        json_path.write_text(__import__('json').dumps(data), encoding='utf-8')
        e = ExcelEngine()
        e.create()
        e.import_json(str(json_path))
        rows = [list(r) for r in e.wb.active.iter_rows(values_only=True)]
        assert rows == [['name', 'age'], ['Alice', 30], ['Bob', 25]]

    def test_import_json_arrays(self, tmp_path: Path) -> None:
        data = [[1, 2], [3, 4]]
        json_path = tmp_path / 'data.json'
        json_path.write_text(__import__('json').dumps(data), encoding='utf-8')
        e = ExcelEngine()
        e.create()
        e.import_json(str(json_path))
        rows = [list(r) for r in e.wb.active.iter_rows(values_only=True)]
        assert rows == [[1, 2], [3, 4]]

    def test_import_json_empty_raises(self, tmp_path: Path) -> None:
        json_path = tmp_path / 'data.json'
        json_path.write_text('[]', encoding='utf-8')
        e = ExcelEngine()
        e.create()
        with pytest.raises(ValueError, match='non-empty'):
            e.import_json(str(json_path))


class TestExcelExtract:
    def test_extract_text(self) -> None:
        e = ExcelEngine()
        e.create()
        ws = e.wb.active
        ws['A1'] = 'x'
        ws['B1'] = 2
        text = e.extract_text()
        assert '[Sheet] x | 2' in text

    def test_extract_tables(self) -> None:
        e = ExcelEngine()
        e.create()
        e.wb.active['A1'] = 'a'
        tables = e.extract_tables()
        assert tables == [[['a']]]

    def test_extract_structure(self) -> None:
        e = ExcelEngine()
        e.create()
        struct = e.extract_structure()
        assert struct['sheets'] == ['Sheet']
