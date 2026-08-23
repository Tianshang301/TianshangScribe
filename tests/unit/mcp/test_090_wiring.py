"""0.9.0 P2 wiring tests: typed ops -> engines, edge actions outside legacy model."""

from __future__ import annotations

from pathlib import Path

from lxml import etree
from openpyxl import Workbook, load_workbook

from tianshang_scribe.core.document import open_document
from tianshang_scribe.mcp.schemas import EditOperation, ToolOptions
from tianshang_scribe.mcp.tools.edit import edit_office_document
from tianshang_scribe.mcp.tools.excel_edit import edit_excel_workbook
from tianshang_scribe.mcp.tools.extract_ppt import extract_presentation_data
from tianshang_scribe.mcp.tools.ppt_create import create_presentation
from tianshang_scribe.mcp.tools.ppt_edit import edit_presentation


def _book(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    ws.append(['k', 'v'])
    ws.append(['a', 1])
    wb.save(str(path))
    return path


def _deck(path: Path) -> Path:
    res = create_presentation(str(path), slides=[{'title': 'T'}])
    assert res['success'] is True, res
    return path


class TestExcelWiring:
    def test_group_rows_then_ungroup(self, tmp_path: Path) -> None:
        book = _book(tmp_path / 'w.xlsx')
        res = edit_excel_workbook(
            str(book), [{'action': 'group_rows', 'range': '2:4', 'hidden': False}]
        )
        assert res['success'] is True, res
        ws = load_workbook(str(book))['Data']
        assert ws.row_dimensions[3].outline_level == 1

        res = edit_excel_workbook(str(book), [{'action': 'ungroup', 'range': '2:4'}])
        assert res['success'] is True, res
        ws = load_workbook(str(book))['Data']
        assert ws.row_dimensions[3].outline_level == 0

    def test_group_columns_with_outline_level(self, tmp_path: Path) -> None:
        book = _book(tmp_path / 'w.xlsx')
        res = edit_excel_workbook(
            str(book),
            [{'action': 'group_columns', 'range': 'B:C', 'outline_level': 2}],
        )
        assert res['success'] is True, res
        ws = load_workbook(str(book))['Data']
        assert ws.column_dimensions['C'].outline_level == 2

    def test_edge_actions_interleaved_preserve_order(self, tmp_path: Path) -> None:
        book = _book(tmp_path / 'w.xlsx')
        res = edit_excel_workbook(
            str(book),
            [
                {'action': 'write_cell', 'cell': 'B2', 'value': 42},
                {'action': 'set_tab_color', 'tab_color': 'FF0000'},
                {'action': 'set_print_area', 'range': 'A1:B2'},
                {
                    'action': 'set_page_setup',
                    'paper_size': 'a4',
                    'orientation': 'landscape',
                },
            ],
        )
        assert res['success'] is True, res
        assert res['data']['total_changes'] == 4
        ws = load_workbook(str(book))['Data']
        assert ws['B2'].value == 42
        tab = ws.sheet_properties.tabColor
        assert tab is not None and str(tab.rgb).endswith('FF0000')
        assert '$A$1' in ws.print_area and '$B$2' in ws.print_area
        assert ws.page_setup.orientation == 'landscape'
        assert ws.page_setup.paperSize == 9  # a4

    def test_edge_action_missing_required_field_errors(self, tmp_path: Path) -> None:
        book = _book(tmp_path / 'w.xlsx')
        res = edit_excel_workbook(str(book), [{'action': 'set_tab_color'}])
        assert res['success'] is False
        assert 'requires' in res['error_message']

    def test_dry_run_shape_unchanged(self, tmp_path: Path) -> None:
        """Legacy dry-run keys stay intact; the validation plan adds more."""
        book = _book(tmp_path / 'w.xlsx')
        res = edit_excel_workbook(
            str(book),
            [{'action': 'group_rows', 'range': '2:4'}],
            options=ToolOptions(dry_run=True),
        )
        assert res['success'] is True, res
        data = res['data']
        # backward-compatible surface
        assert data['dry_run'] is True
        assert data['file'] == str(book)
        assert data['operations'] == 1
        assert data['op_types'] == ['group_rows']
        # enrichment from build_edit_plan
        assert data['all_valid'] is True
        assert data['validations'][0]['ok'] is True
        assert data['estimated_impacted_cells'] == 3


class TestPptWiring:
    def test_set_master_options_and_apply_theme(self, tmp_path: Path) -> None:
        deck = _deck(tmp_path / 'd.pptx')
        res = edit_presentation(
            str(deck),
            [
                {'action': 'set_master_options', 'slide_number': True, 'footer_text': 'Acme'},
                {'action': 'apply_theme', 'theme': 'dark'},
            ],
        )
        assert res['success'] is True, res
        assert res['data']['total_changes'] == 2

        info = extract_presentation_data(str(deck), mode='master_info')
        assert info['success'] is True, info
        layout_types = ' '.join(
            p for m in info['data']['masters'] for lay in m['layouts'] for p in lay['placeholders']
        )
        assert 'SLIDE_NUMBER' in layout_types
        assert 'FOOTER' in layout_types

        ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
        engine = open_document(deck)
        theme_part = engine.prs.slide_masters[0].part.part_related_by(  # type: ignore[union-attr]
            'http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme'
        )
        lt1 = etree.fromstring(theme_part.blob).find('.//a:clrScheme/a:lt1/a:srgbClr', ns)
        assert lt1 is not None and lt1.get('val') == '202124'

    def test_set_master_options_requires_an_option(self, tmp_path: Path) -> None:
        deck = _deck(tmp_path / 'd.pptx')
        res = edit_presentation(str(deck), [{'action': 'set_master_options'}])
        assert res['success'] is False
        assert 'requires at least one' in res['error_message']

    def test_add_media_movie(self, tmp_path: Path) -> None:
        deck = _deck(tmp_path / 'd.pptx')
        media = tmp_path / 'clip.mp4'
        media.write_bytes(b'\x00\x00\x00\x18ftypmp42isom')
        res = edit_presentation(
            str(deck),
            [
                {
                    'action': 'add_media',
                    'slide_index': 0,
                    'media': {'path': str(media), 'kind': 'movie', 'width': 4.0},
                }
            ],
        )
        assert res['success'] is True, res
        engine = open_document(deck)
        kinds = [str(sh.shape_type) for sh in engine.prs.slides[0].shapes]  # type: ignore[union-attr]
        assert any('MEDIA' in k for k in kinds)

    def test_legacy_editoperation_surface(self, tmp_path: Path) -> None:
        deck = _deck(tmp_path / 'd.pptx')
        audio = tmp_path / 'note.mp3'
        audio.write_bytes(b'ID3 ')
        res = edit_office_document(
            str(deck),
            [
                EditOperation(
                    action='add_media', path=str(audio), slide_index=0, media_type='audio'
                ),
                # plain dict: the pydantic Literal rejects unknown actions by
                # design, so direct calls exercise the dispatcher via dict.
                {'action': 'nonexistent_action'},
            ],
        )
        # first op applied, second rejected with INVALID_PARAMETER (1006)
        assert res['success'] is False
        assert res['error_code'] == 1006
        assert 'Unknown edit action' in res['error_message']

        book = _book(tmp_path / 'w.xlsx')
        res = edit_office_document(str(book), [EditOperation(action='group_rows', range='2:4')])
        assert res['success'] is True, res
        ws = load_workbook(str(book))['Data']
        assert ws.row_dimensions[4].outline_level == 1
