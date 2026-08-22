"""Excel document engine built on openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from tianshang_scribe.core.document import DocumentABC
from tianshang_scribe.rendering.styles import TextStyle


class ExcelEngine(DocumentABC):
    """Excel document engine: create, edit, style and convert workbooks."""

    def __init__(self, path: str | Path | None = None) -> None:
        """Initialize the engine with an optional workbook path."""
        super().__init__(path)
        self._wb: Workbook | None = None
        self._base_style: TextStyle = TextStyle.default_excel()
        self._selected_sheet: str | None = None

    @property
    def wb(self) -> Workbook:
        """Return the loaded workbook, raising if none is open."""
        if self._wb is None:
            raise RuntimeError('No workbook loaded. Call create() or open() first.')
        return self._wb

    def create(self) -> None:
        """Create a new blank workbook."""
        self._wb = Workbook()
        self._path = None
        self._base_style = TextStyle.default_excel()
        self._selected_sheet = None

    def open(self, path: str | Path) -> None:
        """Open an existing workbook from the given path."""
        self._path = Path(path)
        if not self._path.exists():
            raise FileNotFoundError(f'File not found: {self._path}')
        self._wb = load_workbook(str(self._path))
        self._base_style = TextStyle.default_excel()
        self._selected_sheet = None

    def select_sheet(self, name: str) -> None:
        """Target subsequent operations at the worksheet named ``name``."""
        if name not in self.wb.sheetnames:
            raise ValueError(f'Sheet not found: {name}')
        self._selected_sheet = name

    def _ws(self) -> Any:
        """Return the targeted worksheet, or the active sheet when none selected."""
        if self._selected_sheet is not None and self._selected_sheet in self.wb.sheetnames:
            return self.wb[self._selected_sheet]
        return self.wb.active

    def save(self, path: str | Path | None = None) -> None:
        """Save the workbook to the given path or the current one."""
        if path is not None:
            self._path = Path(path)
        if self._path is None:
            raise ValueError('No output path specified.')
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(self._path))

    def get_base_style(self) -> TextStyle:
        """Return the engine's base text style."""
        return self._base_style

    def add_text(
        self,
        text: str,
        bold: bool = False,
        italic: bool = False,
        font_name: str | None = None,
        font_size: int | None = None,
        color: str | None = None,
        alignment: str | None = None,
        text_style: TextStyle | None = None,
        **kwargs: Any,
    ) -> Any:
        """Write text into the active sheet and return the last written cell."""
        column = int(kwargs.get('column', 1))
        inline = TextStyle(
            bold=bold or None,
            italic=italic or None,
            font_name=font_name,
            font_size=font_size,
            color=color,
            alignment=alignment,
        )
        final = self._base_style
        if text_style is not None:
            final = final.merge(text_style)
        final = final.merge(inline)

        ws = self._ws()
        if ws.max_row is None or (ws.max_row == 1 and ws['A1'].value is None):
            row = 1
        else:
            row = ws.max_row + 1

        last_cell = None
        for line in text.split('\n'):
            cell = ws.cell(row=row, column=column, value=line)
            font_kwargs: dict[str, Any] = {}
            if final.font_name:
                font_kwargs['name'] = final.font_name
            if final.font_size:
                font_kwargs['size'] = final.font_size
            if final.bold is not None:
                font_kwargs['bold'] = final.bold
            if final.italic is not None:
                font_kwargs['italic'] = final.italic
            if final.color:
                font_kwargs['color'] = final.color.lstrip('#')
            cell.font = Font(**font_kwargs)
            if final.alignment:
                align_map = {
                    'left': 'left',
                    'center': 'center',
                    'right': 'right',
                    'justify': 'justify',
                }
                cell.alignment = Alignment(horizontal=align_map.get(final.alignment, 'left'))
            last_cell = cell
            row += 1

        return last_cell

    def add_styled_content(
        self,
        tokens: list[dict[str, Any]],
    ) -> Any:
        """Write styled content derived from parsed LaTeX-style tokens."""
        current_style = self._base_style
        results: list[Any] = []

        for token in tokens:
            content_kind = token.get('type', 'text')
            content = token.get('content', '')

            if content_kind == 'text':
                results.append(self.add_text(str(content), text_style=current_style))
            elif content_kind == 'command':
                cmd = token.get('command', '')
                if cmd in ('newpage', 'heading', 'includegraphics'):
                    continue
                token_style = TextStyle.from_latex_token(token)
                merged = current_style.merge(token_style)
                inner = token.get('content', '')
                if inner:
                    results.append(self.add_text(inner, text_style=merged))

        return results

    def add_latex_content(self, text: str) -> Any:
        """Parse LaTeX-style markup and add it as styled content."""
        from tianshang_scribe.rendering.latex_parser import parse_structured

        tokens = parse_structured(text)
        return self.add_styled_content(tokens)

    def replace_text(self, old: str, new: str, regex: bool = False) -> int:
        """Replace all occurrences of old text across the workbook; return count."""
        import re as _re

        count = 0
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    if regex:
                        new_val = _re.sub(old, new, str(cell.value))
                        if new_val != str(cell.value):
                            cell.value = new_val
                            count += 1
                    else:
                        if old in str(cell.value):
                            cell.value = str(cell.value).replace(old, new)
                            count += 1
        return count

    def set_style(self, style_str: str) -> None:
        """Merge a style string into the engine's base style."""
        new_style = TextStyle.from_string(style_str)
        self._base_style = self._base_style.merge(new_style)

    def apply_style_to_all(self) -> None:
        """Apply the base style to every populated cell in the workbook."""
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        font_kwargs: dict[str, Any] = {}
                        if self._base_style.font_name:
                            font_kwargs['name'] = self._base_style.font_name
                        if self._base_style.font_size:
                            font_kwargs['size'] = self._base_style.font_size
                        if self._base_style.bold is not None:
                            font_kwargs['bold'] = self._base_style.bold
                        if self._base_style.italic is not None:
                            font_kwargs['italic'] = self._base_style.italic
                        if self._base_style.color:
                            font_kwargs['color'] = self._base_style.color.lstrip('#')
                        if font_kwargs:
                            cell.font = Font(**font_kwargs)

    def to_pdf(self, output_path: str | Path) -> None:
        """Convert the workbook to a PDF at the given output path."""
        self.save()
        from tianshang_scribe.transform.pdf import excel_to_pdf

        excel_to_pdf(str(self._path), str(output_path))

    def add_sheet(self, name: str) -> Any:
        """Create a new worksheet with the given name."""
        return self.wb.create_sheet(title=name)

    def delete_sheet(self, name: str) -> None:
        """Delete the worksheet with the given name if it exists."""
        if name in self.wb.sheetnames:
            del self.wb[name]

    def rename_sheet(self, old_name: str, new_name: str) -> None:
        """Rename an existing worksheet from old_name to new_name."""
        if old_name in self.wb.sheetnames:
            self.wb[old_name].title = new_name

    def set_column_width(self, col_index: int, width: float) -> None:
        """Set the width of the column at the 1-based index in the active sheet."""
        ws = self._ws()
        ws.column_dimensions[get_column_letter(col_index)].width = width

    def set_row_height(self, row_index: int, height: float) -> None:
        """Set the height of the row at the given index in the active sheet."""
        ws = self._ws()
        ws.row_dimensions[row_index].height = height

    def set_formula(self, cell_ref: str, formula: str) -> None:
        """Set a formula on the given cell reference in the active sheet."""
        ws = self._ws()
        ws[cell_ref] = formula

    def import_csv(self, csv_path: str | Path) -> None:
        """Import data from a CSV file into the active sheet."""
        import csv

        ws = self._ws()
        with open(csv_path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row_idx, row_data in enumerate(reader, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

    def import_json(self, json_path: str | Path) -> None:
        """Import a JSON array of objects or lists into the active sheet."""
        import json

        with open(json_path, encoding='utf-8') as f:
            data = json.load(f)

        if not isinstance(data, list) or not data:
            raise ValueError('import_json expects a non-empty JSON array')

        ws = self._ws()
        if all(isinstance(item, dict) for item in data):
            headers = list(dict.fromkeys(k for item in data for k in item))
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            for row_idx, item in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=item.get(header))
        else:
            for row_idx, row_data in enumerate(data, start=1):
                if not isinstance(row_data, list):
                    row_data = [row_data]
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

    def export_csv(self, output_path: str | Path) -> None:
        """Export the active sheet to a CSV file at the given path."""
        import csv

        ws = self._ws()
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

    def get_metadata(self) -> dict[str, str | None]:
        """Return workbook core properties as a metadata dictionary."""
        return {
            'author': self.wb.properties.creator,
            'title': self.wb.properties.title,
            'subject': self.wb.properties.subject,
            'category': self.wb.properties.category,
            'keywords': self.wb.properties.keywords,
            'comments': self.wb.properties.description,
        }

    def set_metadata(self, **kwargs: str) -> None:
        """Set workbook core properties from keyword arguments."""
        props = self.wb.properties
        for key, value in kwargs.items():
            key_lower = key.lower()
            if key_lower == 'author':
                props.creator = value
            elif key_lower == 'title':
                props.title = value
            elif key_lower == 'subject':
                props.subject = value
            elif key_lower == 'category':
                props.category = value
            elif key_lower == 'keywords':
                props.keywords = value
            elif key_lower == 'comments':
                props.description = value

    def add_comment(self, cell_ref: str, text: str) -> None:
        """Attach a comment to the given cell in the active sheet."""
        from openpyxl.comments import Comment

        ws = self._ws()
        ws[cell_ref].comment = Comment(text, 'TianshangScribe')

    def set_protection(self, password: str) -> None:
        """Protect the workbook with the given password."""
        self.wb.security.workbook_password = password

    def unprotect(self) -> None:
        """Remove workbook protection."""
        self.wb.security.workbook_password = ''

    def export_json(self, output_path: str | Path) -> None:
        """Export the active sheet to a JSON array of objects."""
        import json

        ws = self._ws()
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump([], f)
            return
        headers = [str(c) if c is not None else f'col_{i + 1}' for i, c in enumerate(rows[0])]
        data = [
            {headers[j]: str(c) if c is not None else '' for j, c in enumerate(r)} for r in rows[1:]
        ]
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def export_html(self, output_path: str | Path) -> None:
        """Export the active sheet to an HTML table file."""
        ws = self._ws()
        rows = list(ws.iter_rows(values_only=True))
        html_parts = [
            '<!DOCTYPE html><html><head><meta charset="utf-8"><title>',
            self.wb.properties.title or 'Excel Export',
            '</title><style>table{border-collapse:collapse}',
            'td,th{border:1px solid #ccc;padding:4px 8px}</style>',
            '</head><body><table>',
        ]
        first = True
        for row in rows:
            html_parts.append('<tr>')
            for cell in row:
                tag = 'th' if first else 'td'
                html_parts.append(f'<{tag}>{cell or ""}</{tag}>')
            html_parts.append('</tr>')
            first = False
        html_parts.append('</table></body></html>')
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(''.join(html_parts))

    def sort(
        self,
        cell_range: str,
        order: str = 'asc',
        key_columns: list[int] | None = None,
        orders: list[str] | None = None,
    ) -> None:
        """Sort a range of rows in the active sheet.

        The whole row is moved together (column integrity preserved), unlike the
        previous single-column implementation which only reordered one column.
        ``key_columns`` are 0-based column offsets within the range; ``orders``
        is the per-key ``'asc'``/``'desc'`` list. Mixed value types are sorted
        deterministically (numbers < strings < other < None) without raising.
        """
        import re

        from openpyxl.utils import column_index_from_string

        ws = self._ws()
        match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', cell_range)
        if not match:
            raise ValueError(f'Invalid cell range: {cell_range}')
        c1, r1_s, c2, r2_s = match.groups()
        r1 = int(r1_s)
        r2 = int(r2_s)
        col1 = column_index_from_string(c1)
        col2 = column_index_from_string(c2)
        width = col2 - col1 + 1

        keys = key_columns if key_columns is not None else [0]
        key_dirs = orders if orders is not None else [order] * len(keys)
        if len(key_dirs) != len(keys):
            raise ValueError('orders must have the same length as key_columns')

        rows: list[list[Any]] = []
        for r in range(r1, r2 + 1):
            rows.append([ws.cell(row=r, column=c).value for c in range(col1, col2 + 1)])

        def _norm(value: Any) -> tuple[Any, ...]:
            if value is None:
                return (3, 0, '')
            if isinstance(value, bool):
                return (2, int(value), '')
            if isinstance(value, (int, float)):
                return (0, value, '')
            if isinstance(value, str):
                return (1, 0, value.lower())
            return (2, 0, str(value))

        def _make_key(col_index: int) -> Any:
            return lambda r: _norm(r[col_index])

        for k, direction in reversed(list(zip(keys, key_dirs, strict=False))):
            if not (0 <= k < width):
                raise ValueError(f'key_columns index {k} out of range for width {width}')
            rows.sort(key=_make_key(k), reverse=(direction == 'desc'))

        for offset, row in enumerate(rows):
            target = r1 + offset
            for col_offset, value in enumerate(row):
                ws.cell(row=target, column=col1 + col_offset).value = value

    def add_chart(self, chart_type: str, data_range: str, position: str = 'E2') -> None:
        """Add a bar, line or pie chart over the given data range."""
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        ws = self._ws()
        chart_classes = {'bar': BarChart, 'line': LineChart, 'pie': PieChart}
        chart_cls = chart_classes.get(chart_type)
        if chart_cls is None:
            raise ValueError(f'Unsupported chart type: {chart_type}. Use bar, line, or pie.')
        chart = chart_cls()
        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, position)

    def freeze_panes(self, cell: str = 'A2') -> None:
        """Freeze rows above / columns left of ``cell`` (e.g. 'A2', 'B1', 'C4')."""
        self._ws().freeze_panes = cell

    def merge_workbooks(self, paths: list[str]) -> None:
        """Merge the sheets of other workbooks into this one."""
        for p in paths:
            src = load_workbook(p)
            for sheet in src.worksheets:
                new_sheet = self.wb.create_sheet(title=sheet.title)
                for row in sheet.iter_rows():
                    for cell in row:
                        new_sheet[cell.coordinate] = cell.value

    def split_by_sheet(self, output_dir: str | Path) -> list[Path]:
        """Split each worksheet into its own workbook in the output dir."""
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        results: list[Path] = []
        for sheet in self.wb.worksheets:
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet.title
            for row in sheet.iter_rows():
                for cell in row:
                    new_ws[cell.coordinate] = cell.value
            out_path = output_dir / f'{sheet.title}.xlsx'
            new_wb.save(str(out_path))
            results.append(out_path)
        return results

    def clear_content(self) -> None:
        """Clear all cell values in the workbook."""
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

    def clear_formats(self) -> None:
        """Reset fonts, fills, alignments and borders on all cells."""
        from copy import copy

        from openpyxl.styles import Alignment, Border, Font, PatternFill

        empty_font = Font()
        no_fill = PatternFill()
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = copy(empty_font)
                    cell.fill = copy(no_fill)
                    cell.alignment = Alignment()
                    cell.border = Border()

    def clear_links(self) -> None:
        """Remove all hyperlinks from the workbook."""
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.hyperlink = None

    def extract_text(self) -> str:
        """Extract all worksheet content as plain text lines."""
        lines: list[str] = []
        for ws in self.wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = ['' if c is None else str(c) for c in row]
                if any(cells):
                    lines.append(f'[{ws.title}] ' + ' | '.join(cells))
        return '\n'.join(lines)

    def extract_tables(self) -> list[list[list[str]]]:
        """Extract non-empty worksheets as lists of string tables."""
        tables: list[list[list[str]]] = []
        for ws in self.wb.worksheets:
            rows = [
                ['' if c is None else str(c) for c in row]
                for row in ws.iter_rows(values_only=True)
                if any(c is not None for c in row)
            ]
            if rows:
                tables.append(rows)
        return tables

    def extract_images(self, output_dir: str | Path) -> list[Path]:
        """Save embedded workbook images to the output dir and list paths."""
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        saved: list[Path] = []
        idx = 0
        for ws in self.wb.worksheets:
            for image in getattr(ws, '_images', []):
                data = getattr(image, '_data', None)
                if callable(data):
                    try:
                        blob = data()
                    except Exception:  # unreadable image part: skip, do not abort extraction
                        blob = None
                else:
                    blob = data
                if blob is None:
                    continue
                ext = Path(image.format or 'png').suffix or f'.{image.format or "png"}'
                target = out / f'{ws.title}_{idx}{ext}'
                target.write_bytes(blob)
                saved.append(target)
                idx += 1
        return saved

    def extract_structure(self) -> dict[str, Any]:
        """Return workbook structure summary (sheet names and image count)."""
        return {
            'sheets': [ws.title for ws in self.wb.worksheets],
            'images': len([img for ws in self.wb.worksheets for img in getattr(ws, '_images', [])]),
        }
