"""compare_excel_workbooks — Read-only diff between two Excel workbooks.

Compares workbooks without modifying either input. ``structure`` mode reports
sheet-level additions/removals/renames and dimension drift; ``data`` mode
additionally streams a cell-level cached-value diff (with numeric
``tolerance``); ``formula`` mode diffs formula strings only; ``full`` mode
reports every stored-content difference (literals and formulas).
"""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Annotated, Any, Literal

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import Field

from tianshang_scribe.mcp.errors import McpErrorCode, error_response, success_response

#: Per-sheet cap on reported cell diffs before truncation (PLAN.md B.3).
MAX_CELL_DIFFS_PER_SHEET = 10_000

#: Large-workbook guardrails for formula-bearing modes (PLAN.md B.3). Scans
#: still proceed; the result carries an explicit ``warnings`` list instead.
FORMULA_MODE_MAX_BYTES = 50 * 1024 * 1024
FORMULA_MODE_MAX_CELLS = 1_000_000

_MODES = ('structure', 'data', 'formula', 'full')


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith('=')


def _stream_hash(ws: Any) -> str:
    """Stable hash of a sheet's stored-content stream (cheap sheet-level prefilter).

    With ``data_only=True`` the stream is cached values; with
    ``data_only=False`` it is formulas-plus-literals — either way equal hashes
    imply nothing to report for the corresponding mode.
    """
    digest = hashlib.sha256()
    for row in ws.iter_rows(values_only=True):
        digest.update(repr(row).encode('utf-8', 'surrogatepass'))
    return digest.hexdigest()


def _values_equal(a: Any, b: Any, tolerance: float) -> bool:
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) <= tolerance
    return bool(a == b)


def _diff_stream(
    ws_a: Any, ws_b: Any, tolerance: float, cap: int, mode: str
) -> tuple[list[dict[str, Any]], bool]:
    """Stream-compare two sheets position-by-position (O(1) memory).

    ``mode`` selects what is reported: ``'data'`` every cached-value change,
    ``'formula'`` only differences involving at least one formula,
    ``'full'`` everything with a per-cell ``kind`` tag.
    """
    diffs: list[dict[str, Any]] = []
    truncated = False
    rows_a = ws_a.iter_rows(values_only=True)
    rows_b = ws_b.iter_rows(values_only=True)
    max_rows = max(ws_a.max_row or 0, ws_b.max_row or 0)
    for r in range(1, max_rows + 1):
        row_a = next(rows_a, ())
        row_b = next(rows_b, ())
        width = max(len(row_a), len(row_b))
        for c in range(1, width + 1):
            va = row_a[c - 1] if c <= len(row_a) else None
            vb = row_b[c - 1] if c <= len(row_b) else None
            if _values_equal(va, vb, tolerance):
                continue
            has_formula = _is_formula(va) or _is_formula(vb)
            if mode == 'formula' and not has_formula:
                continue
            kind = 'formula' if has_formula and mode != 'data' else 'value'
            diffs.append(
                {'cell': f'{get_column_letter(c)}{r}', 'kind': kind, 'before': va, 'after': vb}
            )
            if len(diffs) >= cap:
                truncated = True
                return diffs, truncated
    return diffs, truncated


def _formula_scan_warnings(path_a: str, path_b: str, wb_a: Any, wb_b: Any) -> list[str]:
    """Large-book advisories for ``formula`` / ``full`` modes (PLAN.md B.3)."""
    warnings: list[str] = []
    for label, p in (('path_a', path_a), ('path_b', path_b)):
        size = Path(p).stat().st_size
        if size > FORMULA_MODE_MAX_BYTES:
            warnings.append(
                f'{label} is {size} bytes (> {FORMULA_MODE_MAX_BYTES}); '
                'the formula scan may be slow.'
            )
    est_cells = sum(
        (ws.max_row or 0) * (ws.max_column or 0)
        for wb in (wb_a, wb_b)
        for ws in (wb[name] for name in wb.sheetnames)
    )
    if est_cells > FORMULA_MODE_MAX_CELLS:
        warnings.append(
            f'workbooks hold ~{est_cells} cells (> {FORMULA_MODE_MAX_CELLS}); '
            'the cell-level scan may be slow.'
        )
    return warnings


def compare_excel_workbooks(
    path_a: Annotated[str, Field(description='First .xlsx workbook.')],
    path_b: Annotated[str, Field(description='Second .xlsx workbook.')],
    mode: Annotated[
        Literal['structure', 'data', 'formula', 'full'],
        Field(
            description=(
                "Diff depth: 'structure' (sheets/dims), 'data' (+cached values), "
                "'formula' (+formula strings only), 'full' (all stored content)."
            )
        ),
    ] = 'structure',
    tolerance: Annotated[
        float, Field(description='Numeric tolerance applied to value comparison (data mode).')
    ] = 0.0,
    sheet_mapping: Annotated[
        dict[str, str] | None,
        Field(
            description="Rename map applied to path_b sheets before diffing, e.g. {'Data2024': 'Data'}."
        ),
    ] = None,
) -> dict[str, Any]:
    """Compare two Excel workbooks and report sheet/cell/formula differences.

    Read-only — never modifies either input file. ``data`` compares cached
    cell values, ``formula`` diffs formula strings only, and ``full`` compares
    all stored content (literals plus formulas). For PowerPoint inspection use
    extract_presentation_data.
    """
    for label, p in (('path_a', path_a), ('path_b', path_b)):
        if not Path(p).exists():
            return error_response(McpErrorCode.DOCUMENT_NOT_FOUND, f"'{p}' not found ({label}).")
        if not str(p).lower().endswith(('.xlsx', '.xlsm')):
            return error_response(
                McpErrorCode.UNSUPPORTED_FORMAT,
                f"'{p}' is not an .xlsx/.xlsm workbook.",
            )

    mapping = dict(sheet_mapping or {})
    try:
        cell_level = mode in ('data', 'formula', 'full')
        data_mode = mode == 'data'
        wb_a = load_workbook(path_a, read_only=True, data_only=data_mode)
        wb_b = load_workbook(path_b, read_only=True, data_only=data_mode)

        mapped_b: dict[str, str] = {mapping.get(n, n): n for n in wb_b.sheetnames}
        names_a = set(wb_a.sheetnames)
        added = sorted(set(mapped_b) - names_a)
        removed = sorted(names_a - set(mapped_b))
        renamed = [
            {'from': orig, 'to': eff} for eff, orig in sorted(mapped_b.items()) if orig != eff
        ]
        common = sorted(names_a & set(mapped_b))

        sheets_block: dict[str, Any] = {
            'added': added,
            'removed': removed,
            'renamed': renamed,
        }
        warnings: list[str] = []
        if mode in ('formula', 'full'):
            warnings = _formula_scan_warnings(path_a, path_b, wb_a, wb_b)

        cells: list[dict[str, Any]] = []
        truncated = False
        identical: list[str] = []
        dims_changed: list[dict[str, Any]] = []

        for name in common:
            ws_a = wb_a[name]
            ws_b = wb_b[mapped_b[name]]
            if cell_level and _stream_hash(ws_a) == _stream_hash(ws_b):
                identical.append(name)
                continue
            if mode == 'structure':
                ra, ca = ws_a.max_row or 0, ws_a.max_column or 0
                rb, cb = ws_b.max_row or 0, ws_b.max_column or 0
                if (ra, ca) != (rb, cb):
                    dims_changed.append(
                        {
                            'sheet': name,
                            'a': {'rows': ra, 'cols': ca},
                            'b': {'rows': rb, 'cols': cb},
                        }
                    )
                continue
            sheet_diffs, hit_cap = _diff_stream(
                ws_a, ws_b, tolerance, MAX_CELL_DIFFS_PER_SHEET, mode
            )
            truncated = truncated or hit_cap
            for d in sheet_diffs:
                cells.append({'sheet': name, **d})

        wb_a.close()
        wb_b.close()
        payload: dict[str, Any] = {
            'path_a': path_a,
            'path_b': path_b,
            'mode': mode,
            'tolerance': tolerance,
            'sheets': sheets_block,
        }
        if mode == 'structure':
            payload['dims_changed'] = dims_changed
        else:
            payload['cells'] = cells
            payload['cell_diff_count'] = len(cells)
            payload['truncated'] = truncated
            payload['identical_sheets'] = identical
        if warnings:
            payload['warnings'] = warnings
        return success_response(payload)
    except Exception as e:  # surface any read failure as a structured error
        return error_response(McpErrorCode.INTERNAL_ERROR, str(e))
